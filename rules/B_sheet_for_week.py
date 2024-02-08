from tools import get_or_create_money_style, log_message
import streamlit as st
from openpyxl.utils import get_column_letter


def create_sheet_for_week(workbook, sheet):
    extracted_data = st.session_state.extracted_data
    weeks_info = st.session_state.weeks_info

    for idx, week_info in enumerate(weeks_info, start=1):
        week_name = "week {}".format(idx)
        new_sheet = workbook.create_sheet(week_name)

        for col in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col)
            new_sheet.column_dimensions[col_letter].width = 15

        week_title = week_info["week_title"]
        create_table(new_sheet, week_title, extracted_data[week_name])

        money_style = get_or_create_money_style(workbook)

        for col in ["B", "E", "G", "I", "K", "M", "O", "Q", "R"]:
            for cell in new_sheet[col]:
                cell.style = money_style
                cell.number_format = "$#,##0"


def create_table(sheet, week_title, week_data):
    sheet.append([week_title])
    sheet.append(
        [
            "Colaborador",
            "Salario",
            "Hr laboradas",
            "Hr Rec nocturnas",
            "Rec pago nocturnas",
            "Hr Rec dia Dom-Fest ",
            "Rec pago dia Dom-Fest ",
            "Hr Rec Noche Dom-fes",
            "Rec pago noche Dom-fest",
            "Hr extra dia",
            "Hr pago extra dia",
            "Hr extra Noche",
            "Hr pago extra noche",
            "Hr extra día Dom-fes",
            "Hr pago extra día Dom-fes",
            "Hr extra noche Dom-fes",
            "Hr pago extra noche Dom-fes",
            "Total Sem",
        ]
    )

    for employe_name, employe_info in week_data.items():
        salary = employe_info["salary"]
        total_hours = employe_info["total_hours"]
        night_surchage_hours = employe_info["night_surchage_hours"]
        night_surchage_pay = employe_info["night_surchage_pay"]
        day_holidays_surcharges_hours = employe_info["day_holidays_surcharges_hours"]
        day_holidays_surchage_pay = employe_info["day_holidays_surchage_pay"]
        night_holidays_surcharges_hours = employe_info["night_holidays_surcharges_hours"]
        night_holidays_surchage_pay = employe_info["night_holidays_surchage_pay"]
        daytime_overtime_hours = employe_info["daytime_overtime_hours"]
        daytime_overtime_pay = employe_info["daytime_overtime_pay"]
        night_overtime_hours = employe_info["night_overtime_hours"]
        night_overtime_pay = employe_info["night_overtime_pay"]
        daytime_holidays_hours = employe_info["daytime_holidays_hours"]
        daytime_holidays_pay = employe_info["daytime_holidays_pay"]
        night_holidays_overtime_hours = employe_info["night_holidays_overtime_hours"]
        night_holidays_overtime_pay = employe_info["night_holidays_overtime_pay"]

        E = get_column_letter(sheet["E"][0].column)
        G = get_column_letter(sheet["G"][0].column)
        I = get_column_letter(sheet["I"][0].column)  # noqa: E741
        K = get_column_letter(sheet["K"][0].column)
        M = get_column_letter(sheet["M"][0].column)
        O = get_column_letter(sheet["O"][0].column)  # noqa: E741
        Q = get_column_letter(sheet["Q"][0].column)

        row_start = 1

        total_sem = f"={E}{sheet.max_row+row_start}+{G}{sheet.max_row+row_start}+{I}{sheet.max_row+row_start}+{K}{sheet.max_row+row_start}+{M}{sheet.max_row+row_start}+{O}{sheet.max_row+row_start}+{Q}{sheet.max_row+row_start}"

        sheet.append(
            [
                employe_name,
                salary,
                total_hours,
                night_surchage_hours,
                night_surchage_pay,
                day_holidays_surcharges_hours,
                day_holidays_surchage_pay,
                night_holidays_surcharges_hours,
                night_holidays_surchage_pay,
                daytime_overtime_hours,
                daytime_overtime_pay,
                night_overtime_hours,
                night_overtime_pay,
                daytime_holidays_hours,
                daytime_holidays_pay,
                night_holidays_overtime_hours,
                night_holidays_overtime_pay,
                total_sem,
            ]
        )


def main(workbook, progress_bar):
    for sheet in workbook:
        create_sheet_for_week(workbook, sheet)
        break
    log_message("the sheets per week and their respective tables were created.")
    if progress_bar is not None:
        progress_bar.progress(1.0)

    return workbook
