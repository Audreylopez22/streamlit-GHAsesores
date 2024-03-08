from tools import log_message
import streamlit as st
import io
from openpyxl import load_workbook
import pandas as pd
import os
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import re

st.set_page_config(page_title="PDF Download", page_icon="📄", layout="wide")

st.markdown("# PDF Download")
st.sidebar.header("PDF Download")

if (
    "authentication_status" not in st.session_state
    or st.session_state.authentication_status is None
    or st.session_state.authentication_status is False
):
    st.warning("You must login to access this page.")
    st.markdown(
        f'<meta http-equiv="refresh" content="0;url={st.secrets.urls.login}">',
        unsafe_allow_html=True,
    )
    st.stop()


def date(month):
    pattern = r"\b\d+\b"
    data = st.session_state.weeks_info
    year = st.session_state.period

    first_week = data[0]
    numbers = re.findall(pattern, first_week["week_title"])
    start_date = numbers[0]
    first_week["start_date"] = start_date

    last_week = data[-1]
    numbers = re.findall(pattern, last_week["week_title"])
    end_date = numbers[1]
    last_week["end_date"] = end_date

    year_number = re.sub(r"\D", "", year)

    if int(start_date) >= int(end_date):
        full_start_date = f"{start_date}/{month}/{year_number}"

        if month == "12":
            next_month = 1
            next_year = int(year_number) + 1
        else:
            next_month = int(month) + 1
            next_year = year_number
        full_end_date = f"{end_date}/{next_month:02}/{next_year}"
    else:
        full_start_date = f"{start_date}/{month}/{year_number}"
        full_end_date = f"{end_date}/{month}/{year_number}"

    return full_start_date, full_end_date


def filter_and_display_data(sheet):
    if sheet.title == "SIMPLE SHEET":
        log_message(f"Filtering and displaying data for sheet: {sheet.title}")
        header_row = sheet[1]
        df = pd.DataFrame(
            sheet.iter_rows(min_row=2, values_only=True),
            columns=[cell.value for cell in header_row],
        )
        return df

    else:
        return pd.DataFrame()


def generar_pdf(data_row, pdf_path, full_start_date, full_end_date):
    image = os.path.join("reports", "Volante.jpg")

    try:
        width, height = letter
        with open(pdf_path, "wb") as output_file:
            c = canvas.Canvas(output_file)
            c.setPageSize(letter)

            c.drawImage(image, x=0, y=0, width=width, height=height)
            c.setFontSize(8)
            # date
            x = 293
            y = 682
            start_date = str(full_start_date)
            c.drawString(x, y, start_date)
            x = 400
            y = 682
            end_date = str(full_end_date)
            c.drawString(x, y, end_date)

            # Colaborador
            x = 150
            y = 666
            name = data_row["Colaborador"]
            c.drawString(x, y, name)
            # salary
            x = 150
            y = 650
            salary = "${:,.2f}".format(float(str(data_row["Salario"])))
            c.drawString(x, y, salary)
            # Recargo Noctuno
            # hours
            x = 370
            y = 622
            hr_rec_night = str(data_row["Hr Rec nocturnas"])
            c.drawString(x, y, hr_rec_night)
            # pago
            x = 470
            y = 622
            pay_rec_night = "${:,.2f}".format(float(str(data_row["Rec pago nocturnas"])))
            c.drawString(x, y, pay_rec_night)
            # Horas Extra diurnas
            # hours
            x = 370
            y = 607
            hr_extra_day = str(data_row["Hr extra dia"])
            c.drawString(x, y, hr_extra_day)
            # pago
            x = 470
            y = 607
            pay_extra_day = "${:,.2f}".format(float(str(data_row["Hr pago extra dia"])))
            c.drawString(x, y, pay_extra_day)
            # Horas nocturnas
            # hours
            x = 370
            y = 591
            hr_extra_night = str(data_row["Hr extra Noche"])
            c.drawString(x, y, hr_extra_night)
            # pago
            x = 470
            y = 591
            pay_extra_night = "${:,.2f}".format(float(str(data_row["Hr pago extra noche"])))
            c.drawString(x, y, pay_extra_night)

            # Horas dia dom-fes
            # hours
            x = 370
            y = 575
            hr_extra_day_holidays = str(data_row["Hr extra día Dom-fes"])
            c.drawString(x, y, hr_extra_day_holidays)
            # pago
            x = 470
            y = 575
            pay_extra_day_holidays = "${:,.2f}".format(
                float(str(data_row["Hr pago extra día Dom-fes"]))
            )
            c.drawString(x, y, pay_extra_day_holidays)

            # Horas noche dom-fes
            # hours
            x = 370
            y = 559
            hr_extra_night_holidays = str(data_row["Hr extra noche Dom-fes"])
            c.drawString(x, y, hr_extra_night_holidays)
            # pago
            x = 470
            y = 559
            pay_extra_night_holidays = "${:,.2f}".format(
                float(str(data_row["Hr pago extra noche Dom-fes"]))
            )
            c.drawString(x, y, pay_extra_night_holidays)

            # recargo nocturno fest
            # hours
            x = 370
            y = 543
            hr_rec_night_holidays = str(data_row["Hr Rec Noche Dom-fes"])
            c.drawString(x, y, hr_rec_night_holidays)
            # pago
            x = 470
            y = 543
            pay_rec_night_holidays = "${:,.2f}".format(
                float(str(data_row["Rec pago noche Dom-fest"]))
            )
            c.drawString(x, y, pay_rec_night_holidays)

            # recargo dia fest
            # hours
            x = 370
            y = 527
            hr_rec_day_holidays = str(data_row["Hr Rec dia Dom-Fest "])
            c.drawString(x, y, hr_rec_day_holidays)
            # pago
            x = 470
            y = 527
            pay_rec_day_holidays = "${:,.2f}".format(float(str(data_row["Rec pago dia Dom-Fest "])))
            c.drawString(x, y, pay_rec_day_holidays)

            # Total
            # pago
            x = 470
            y = 513
            pay_total = "${:,.2f}".format(float(str(data_row["Total Sem"])))
            c.drawString(x, y, pay_total)

            c.showPage()
            c.save()
        log_message(f"The PDF was generated for  {data_row['Colaborador']}")
    except Exception as e:
        st.error(f"Error generating the PDF: {str(e)}")


def main():
    if "tmp_file" not in st.session_state or st.session_state["period"] == "DEFAULT":
        st.warning("Cannot display data because no file has been uploaded.")
        return
    # st.write(st.session_state.tmp_file)
    directory, file_name = os.path.split(st.session_state.tmp_file)
    file_name_uppercase = file_name.upper()
    uppercased_file = os.path.join(directory, file_name_uppercase)
    # st.write(uppercased_file)

    if os.path.exists(uppercased_file):
        with open(uppercased_file, "rb") as file_content:
            st.session_state.tmp_file_content = file_content.read()

    st.write(st.session_state.tmp_file_content)
    uploaded_file_contents = st.session_state.tmp_file_content

    workbook = load_workbook(io.BytesIO(uploaded_file_contents))

    filtered_data = filter_and_display_data(workbook["SIMPLE SHEET"])

    months = {
        "ENERO": "01",
        "FEBRERO": "02",
        "MARZO": "03",
        "ABRIL": "04",
        "MAYO": "05",
        "JUNIO": "06",
        "JULIO": "07",
        "AGOSTO": "08",
        "SEPTIEMBRE": "09",
        "OCTUBRE": "10",
        "NOVIEMBRE": "11",
        "DICIEMBRE": "12",
    }
    for sheet_name in workbook.sheetnames:
        if sheet_name in months:
            # st.write(f"El nombre de la hoja '{sheet_name}' coincide con un mes.")
            # st.write(f"El número correspondiente es: {months[sheet_name.upper()]}")
            month = months[sheet_name.upper()]
            full_start_date, full_end_date = date(month)
            break
        else:
            st.write(f"The sheet name '{sheet_name}' does not match with any month.")

    st.write("download the pdf by employee :")

    for index, row in filtered_data.iterrows():
        colaborador = row["Colaborador"]
        pdf_filename = f"{colaborador}_reporte.pdf"
        my_path = os.path.join("/tmp", pdf_filename)
        generar_pdf(row, my_path, full_start_date, full_end_date)
        pdf_download_path = f"/tmp/{colaborador}_reporte.pdf"

        st.download_button(
            label=f"Download PDF of {colaborador}",
            data=open(pdf_download_path, "rb"),
            file_name=pdf_filename,
            mime="application/pdf",
        )


if __name__ == "__main__":
    main()
