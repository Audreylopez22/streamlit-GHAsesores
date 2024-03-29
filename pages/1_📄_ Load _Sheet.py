import streamlit as st
import openpyxl
from tools import log_message
import os
from io import BytesIO
import importlib
import formulas
import configparser
from tempfile import NamedTemporaryFile


st.set_page_config(page_title="Load Sheet", page_icon="📑", layout="wide")

if (
    "authentication_status" not in st.session_state
    or st.session_state.authentication_status is None
    or st.session_state.authentication_status is False
):
    st.warning("You must login to access this page.")
    st.markdown(
        f'<meta http-equiv="refresh" content="0;url={st.secrets.urls.login}">',
        unsafe_allow_html=True,
    )
    st.stop()

welcome = "Welcome to GH Asesores payroll Processor!"

st.markdown(f"# {welcome}")
st.sidebar.header(welcome)
st.session_state.tmp_file = None

st.write(
    """ Here you can upload your Excel file to perform various operations. Start by selecting your file using the 'Upload Excel File' button. The application will automatically check if the file contains the necessary data. If so, it will proceed to calculate overtime hours and surcharges for each employee, depending on the number of weeks included in the file."""
)


def load_data_from_excel(uploaded_file):
    workbook = openpyxl.load_workbook(uploaded_file)
    sheet = workbook.active

    data = [[cell.value for cell in row] for row in sheet.iter_rows()]
    return data


def process_rules(workbook, progress_bar):
    try:
        rule_files = [
            "A_hours_per_week.py",
            "B_sheet_for_week.py",
            "C_create_simple_sheet.py",
            "D_process_simple_sheet.py",
        ]
        total_steps = len(rule_files)

        for i, filename in enumerate(rule_files):
            rule_module_name = f"rules.{filename[:-3]}"
            rule = importlib.import_module(rule_module_name)
            if hasattr(rule, "main") and callable(rule.main):
                workbook = rule.main(workbook, progress_bar)

            # For the progres bar
            progress_percentage = min(1.0, (i + 1) / total_steps)
            progress_bar.progress(progress_percentage)

    except Exception as error:
        print(error)

    return workbook


def main():
    del st.session_state.tmp_file
    config = configparser.ConfigParser()
    config.read("config.ini")
    # crear st stare de config y de la key que seleccione en el dropdown
    st.session_state["config"] = config

    config_keys = [""]

    for key in config.keys():
        if key == "DEFAULT":
            continue
        else:
            config_keys.append(key)

    period = st.selectbox(
        "please select the reference period to make the calculation. ", config_keys
    )

    st.session_state["period"] = period

    st.write("You selected:", period)

    if os.path.exists("/tmp"):
        for file in os.listdir("/tmp"):
            if file.endswith((".xlsx", ".XLSX", ".pdf", ".PDF")):
                os.unlink(os.path.join(os.sep, "tmp", file))

    uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx", "xls"])

    if uploaded_file is not None:
        uploaded_file_contents = uploaded_file.read()

        if "file_hash" not in st.session_state or st.session_state.file_hash != hash(
            uploaded_file_contents
        ):
            st.session_state.file_hash = hash(uploaded_file_contents)

        st.session_state.data = load_data_from_excel(uploaded_file)

        # st.write(st.session_state.data)

        workbook = openpyxl.load_workbook(uploaded_file)
        sheet = workbook.active

        if (
            sheet.max_row == 1
            and sheet.max_column == 1
            and sheet.cell(row=1, column=1).value is None
        ):
            st.warning("The Excel file is empty.")
            log_message("The Excel file is empty.")

        else:
            progress_bar = st.progress(0.0)
            log_message("Reading the entered file")
            log_message("Making modifications to the Excel file...")

            workbook = process_rules(workbook, progress_bar)

            # the modified or processed file is saved in memory
            modified_file = BytesIO()
            log_message("Saving modifications to the Excel file...")
            workbook.save(modified_file)

            with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                st.session_state.tmp_file = tmp_file.name
                workbook.save(tmp_file.name)

            log_message("Loading...")
            xl_model = formulas.ExcelModel().loads(st.session_state.tmp_file).finish()
            xl_model.calculate()
            xl_model.write(dirpath="/tmp")

            # download button for the downloaded file that is in memory
            st.session_state.download_button = st.download_button(
                label="Download modified file",
                data=modified_file.getvalue(),
                key="download_file.xlsx",
                file_name="modified_file.xlsx",
            )


if __name__ == "__main__":
    main()
