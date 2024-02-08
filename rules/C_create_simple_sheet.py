from tools import log_message
from openpyxl.utils import get_column_letter


def create_simple_sheet(workbook):
    if "Simple Sheet" not in workbook.sheetnames:
        log_message("creating simple sheet")
        simple_sheet = workbook.create_sheet("Simple Sheet")

        new_columns = [
            "Colaborador",
            "Salario",
            "Hr laboradas",
            "Hr Rec nocturnas",
            "Rec pago nocturnas",
            "Hr Rec dia Dom-Fest ",
            "Rec pago dia Dom-Fest ",
            "Hr Rec Noche Dom-fes",
            "Rec pago noche Dom-fest",
            "Hr extra dia",
            "Hr pago extra dia",
            "Hr extra Noche",
            "Hr pago extra noche",
            "Hr extra día Dom-fes",
            "Hr pago extra día Dom-fes",
            "Hr extra noche Dom-fes",
            "Hr pago extra noche Dom-fes",
            "Total Sem",
        ]

        for i, column_name in enumerate(new_columns, start=1):
            new_column_letter = get_column_letter(i)
            simple_sheet[f"{new_column_letter}1"] = column_name

        for col in range(1, simple_sheet.max_column + 1):
            col_letter = get_column_letter(col)
            simple_sheet.column_dimensions[col_letter].width = 15


def main(workbook, progress_bar):
    create_simple_sheet(workbook)

    if progress_bar is not None:
        progress_bar.progress(1.0)

    return workbook
